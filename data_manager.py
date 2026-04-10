"""
data_manager.py  –  BudgetPal
Unified account model: every expense/goal is an account building toward a target.
"""

import os, re
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = os.path.join(os.path.dirname(__file__), "user_data")
os.makedirs(DATA_DIR, exist_ok=True)

CURRENCIES  = {"USD": "$", "JPY": "¥"}
PAY_PERIODS = {"Weekly": 52, "Biweekly": 26, "Monthly": 12}
DAYS_OF_WEEK  = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
DAYS_OF_MONTH = [str(i) for i in range(1, 29)] + ["Last day"]
RECUR_OPTIONS = ["Weekly","Biweekly","Monthly","Quarterly","Annually","One-time"]

# ── Named colour palettes ─────────────────────────────────────────────────────
PALETTES = {
    "White": {   # Light mode — clean minimal (default)
        "dark":   "#2E4057",   # headers, topbar
        "teal":   "#048A81",   # accent buttons
        "light":  "#EAF4FB",   # panel backgrounds
        "gold":   "#F6AE2D",   # highlights / allowance
        "white":  "#FFFFFF",   # main background
        "red":    "#C0392B",   # alerts
        "green":  "#27AE60",   # positive / success
        "gray":   "#BDC3C7",   # subtle elements
        "purple": "#8E44AD",   # running expenses
        "font":   "Arial",
    },
    "Black": {   # Dark mode
        "dark":   "#121212",   # headers, topbar
        "teal":   "#03DAC6",   # accent
        "light":  "#1E1E1E",   # panel backgrounds
        "gold":   "#FFD700",   # highlights
        "white":  "#2C2C2C",   # main background (dark)
        "red":    "#CF6679",   # alerts
        "green":  "#4CAF50",   # positive
        "gray":   "#555555",   # subtle
        "purple": "#BB86FC",   # running expenses
        "font":   "Arial",
    },
    "Pink": {
        "dark":   "#7B1048",   # deep rose — headers
        "teal":   "#E91E8C",   # hot pink — accent
        "light":  "#FFF0F5",   # blush — panels
        "gold":   "#F48FB1",   # light pink — highlights
        "white":  "#FFF5F8",   # soft white — background
        "red":    "#C62828",   # alerts
        "green":  "#AD1457",   # positive (dark rose)
        "gray":   "#F8BBD0",   # subtle pink
        "purple": "#CE93D8",   # running expenses
        "font":   "Arial",
    },
    "Blue": {
        "dark":   "#0D2B5E",   # navy — headers
        "teal":   "#1565C0",   # royal blue — accent
        "light":  "#E3F2FD",   # sky blue — panels
        "gold":   "#FFA726",   # amber — highlights
        "white":  "#F0F7FF",   # pale blue — background
        "red":    "#D32F2F",   # alerts
        "green":  "#0288D1",   # positive (bright blue)
        "gray":   "#90CAF9",   # subtle blue
        "purple": "#7986CB",   # running expenses
        "font":   "Arial",
    },
    "Red": {
        "dark":   "#7F0000",   # deep crimson — headers
        "teal":   "#D32F2F",   # red — accent
        "light":  "#FFF3F3",   # blush — panels
        "gold":   "#FF8F00",   # amber — highlights
        "white":  "#FFF8F8",   # warm white — background
        "red":    "#B71C1C",   # alerts (darker red)
        "green":  "#558B2F",   # positive (contrast green)
        "gray":   "#FFCDD2",   # subtle rose
        "purple": "#E040FB",   # running expenses
        "font":   "Arial",
    },
}

DEFAULT_THEME = PALETTES["White"]
PALETTE_NAMES = list(PALETTES.keys())

# ── XLSX style constants ──────────────────────────────────────────────────────
THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NORMAL_FONT    = Font(name="Arial", size=10)
HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Arial")
HEADER_FILL    = PatternFill("solid", start_color="2E4057")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
SUBHEADER_FILL = PatternFill("solid", start_color="048A81")
INPUT_FILL     = PatternFill("solid", start_color="EAF4FB")
GOAL_FILL      = PatternFill("solid", start_color="FFF3CD")
ACCT_FILL      = PatternFill("solid", start_color="E8F5E9")

def _profile_path(uid): return os.path.join(DATA_DIR, f"{uid}.xlsx")
def profile_exists(uid): return os.path.exists(_profile_path(uid))
def make_uid(name):
    base = re.sub(r"[^a-z0-9]", "", name.lower())[:20]
    return base if base else "user"
def uid_taken(uid): return profile_exists(uid)

def _style(cell, font=None, fill=None, align=None, border=None, num_fmt=None):
    if font:    cell.font = font
    if fill:    cell.fill = fill
    if align:   cell.alignment = align
    if border:  cell.border = border
    if num_fmt: cell.number_format = num_fmt

def _hdr(cell, value):
    cell.value = value
    _style(cell, font=HEADER_FONT, fill=HEADER_FILL,
           align=Alignment(horizontal="center", vertical="center"), border=BORDER)

def _sub(cell, value):
    cell.value = value
    _style(cell, font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
           align=Alignment(horizontal="left", vertical="center"), border=BORDER)

def _inp(cell, value, num_fmt=None):
    cell.value = value
    _style(cell, font=Font(color="0000FF", name="Arial", size=10),
           fill=INPUT_FILL, border=BORDER, num_fmt=num_fmt)

def _nrm(cell, value, num_fmt=None):
    cell.value = value
    _style(cell, font=NORMAL_FONT, border=BORDER, num_fmt=num_fmt)

def _currency_fmt(currency):
    return '"$"#,##0.00' if currency == "USD" else '"¥"#,##0'

def _ordinal(val):
    """'1' -> '1st', '22' -> '22nd', 'Last day' -> 'last day of month'"""
    if val == "Last day": return "last day of month"
    n = int(val)
    sfx = "th" if 11 <= n <= 13 else {1:"st",2:"nd",3:"rd"}.get(n%10,"th")
    return f"{n}{sfx} of month"

def due_label_from(due_type, due_value):
    if not due_type:
        return "Each paycheck"
    if due_type == "Date":
        return f"Deadline: {due_value}"
    if due_type == "Day of month":
        return f"Due: {_ordinal(due_value)}"
    return f"Due: every {due_value}"


# ══════════════════════════════════════════════════════════════════════════════
#  DATA STRUCTURES
#
#  account  = {
#    "name": str,
#    "balance": float,          # current actual balance
#    "type": "bank",            # pure holding account — no target
#  }
#
#  obligation = {               # expense OR goal — everything is an obligation
#    "name": str,
#    "target": float,           # amount needed by due date
#    "recurrence": str,         # Weekly/Monthly/.../One-time
#    "due_type": str,           # "Day of month" | "Day of week"
#    "due_value": str,          # "1", "15", "Last day", "Saturday", ...
#    "due_date": str,           # ISO date of NEXT due date  YYYY-MM-DD
#    "deposit_account": str,    # name of linked bank account
#    "kind": str,               # "expense" | "goal" | "running" | "allowance"
#    "paid": bool,              # True if awaiting user confirmation this period
#  }
# ══════════════════════════════════════════════════════════════════════════════

# ── Date helpers ──────────────────────────────────────────────────────────────

def _next_due_date(due_type, due_value, recurrence, after=None):
    """
    Return the next due date (date object) on or after `after` (default today).
    For recurrence="One-time" just return the stored date unchanged.
    """
    today = after or date.today()

    if due_type == "Day of month":
        if due_value == "Last day":
            # Last day of current month
            import calendar
            last = calendar.monthrange(today.year, today.month)[1]
            candidate = date(today.year, today.month, last)
        else:
            day = int(due_value)
            try:
                candidate = date(today.year, today.month, day)
            except ValueError:
                candidate = date(today.year, today.month, 28)
        if candidate < today:
            # Roll to next month
            m = today.month + 1 if today.month < 12 else 1
            y = today.year if today.month < 12 else today.year + 1
            if due_value == "Last day":
                import calendar
                last = calendar.monthrange(y, m)[1]
                candidate = date(y, m, last)
            else:
                try:
                    candidate = date(y, m, int(due_value))
                except ValueError:
                    candidate = date(y, m, 28)
        return candidate

    elif due_type == "Day of week":
        target_dow = DAYS_OF_WEEK.index(due_value)
        days_ahead = (target_dow - today.weekday()) % 7
        if days_ahead == 0:
            days_ahead = 7
        return today + timedelta(days=days_ahead)

    return today  # fallback


def advance_due_date(obligation):
    """
    Roll an obligation's due_date forward by one recurrence period.
    Returns updated obligation dict (does not mutate in place).
    """
    ob = dict(obligation)
    if ob["recurrence"] == "One-time":
        return ob  # don't roll one-time obligations

    current = datetime.strptime(ob["due_date"], "%Y-%m-%d").date()
    deltas = {
        "Weekly": timedelta(weeks=1),
        "Biweekly": timedelta(weeks=2),
        "Monthly": None,  # special
        "Quarterly": None,
        "Annually": None,
    }
    rec = ob["recurrence"]
    if rec == "Weekly":
        next_d = current + timedelta(weeks=1)
    elif rec == "Biweekly":
        next_d = current + timedelta(weeks=2)
    elif rec == "Monthly":
        m = current.month + 1 if current.month < 12 else 1
        y = current.year if current.month < 12 else current.year + 1
        import calendar
        if ob["due_value"] == "Last day":
            last = calendar.monthrange(y, m)[1]
            next_d = date(y, m, last)
        else:
            day = min(int(ob["due_value"]), calendar.monthrange(y, m)[1])
            next_d = date(y, m, day)
    elif rec == "Quarterly":
        m = current.month + 3
        y = current.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        import calendar
        day = min(int(ob.get("due_value","1") if ob.get("due_value","1") != "Last day" else "28"),
                  calendar.monthrange(y, m)[1])
        next_d = date(y, m, day)
    elif rec == "Annually":
        try:
            next_d = date(current.year + 1, current.month, current.day)
        except ValueError:
            next_d = date(current.year + 1, current.month, 28)
    else:
        next_d = current

    ob["due_date"] = next_d.isoformat()
    ob["paid"] = False
    return ob


def paychecks_until(due_date_str, pay_period):
    """Number of paychecks between today and due_date (min 1)."""
    try:
        due = datetime.strptime(due_date_str, "%Y-%m-%d").date()
        days_left = max((due - date.today()).days, 1)
        annual = PAY_PERIODS[pay_period]
        days_per_paycheck = 365 / annual
        return max(round(days_left / days_per_paycheck), 1)
    except Exception:
        return 1


# ══════════════════════════════════════════════════════════════════════════════
#  BUDGET ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def build_budget(profile, net_pay):
    """
    Core budget calculation.

    For each obligation:
      deposit_needed = max(target - account_balance, 0) / paychecks_until_due

    If sum(deposits) <= net_pay:
      leftover is spread proportionally back to all obligations (they get ahead).
    If sum(deposits) > net_pay:
      each deposit is scaled down proportionally.

    Returns list of line items and summary.
    """
    pp        = profile["pay_period"]
    accounts  = {a["name"]: a["balance"] for a in profile.get("accounts", [])}
    obligs    = profile.get("obligations", [])

    lines = []
    for ob in obligs:
        acct_bal = accounts.get(ob["deposit_account"], 0.0)
        kind     = ob.get("kind", "expense")

        if kind == "allowance":
            # Allowance: fixed dollar amount every paycheck, no due date.
            # target field holds the desired amount per paycheck.
            needed    = ob["target"]
            goal_rate = ob["target"]   # the full amount is the "rate"
            pc_left   = 1
        else:
            remaining = max(ob["target"] - acct_bal, 0.0)
            pc_left   = paychecks_until(ob["due_date"], pp)
            needed    = remaining / pc_left
            goal_rate = needed  # minimum needed this period

        lines.append({
            "name":            ob["name"],
            "kind":            kind,
            "deposit_account": ob["deposit_account"],
            "target":          ob["target"],
            "due_date":        ob.get("due_date", ""),
            "recurrence":      ob.get("recurrence", ""),
            "acct_balance":    acct_bal,
            "needed":          needed,
            "deposit":         needed,   # may be scaled below
            "goal_rate":       goal_rate,
            "spendable":       0.0,      # computed after scaling
            "pc_left":         pc_left,
        })

    total_needed = sum(l["needed"] for l in lines)

    if total_needed == 0:
        scale = 1.0
        leftover = net_pay
    elif total_needed <= net_pay:
        scale    = 1.0
        leftover = net_pay - total_needed
        # Spread leftover proportionally
        if leftover > 0 and total_needed > 0:
            for l in lines:
                share = (l["needed"] / total_needed) * leftover
                l["deposit"] = l["needed"] + share
        else:
            leftover = net_pay - total_needed
    else:
        scale    = net_pay / total_needed
        leftover = 0.0
        for l in lines:
            l["deposit"] = l["needed"] * scale

    # Compute spendable now that final deposit amounts are known
    # allowance: fully spendable (that's the point)
    # others: max(actual_deposit - goal_rate, 0)
    for l in lines:
        if l["kind"] == "allowance":
            l["spendable"] = l["deposit"]
        else:
            l["spendable"] = max(l["deposit"] - l["goal_rate"], 0.0)
        l["pct"] = (l["deposit"] / net_pay * 100) if net_pay else 0

    return {
        "lines":        lines,
        "net_pay":      net_pay,
        "total_needed": total_needed,
        "scale":        scale,
        "leftover_raw": max(net_pay - total_needed, 0),
        "shortfall":    max(total_needed - net_pay, 0),
    }


def apply_deposits(profile, budget):
    """
    After user confirms paycheck, add deposit amounts to each linked account.
    Returns updated profile (does not mutate in place).
    """
    import copy
    p = copy.deepcopy(profile)
    acct_map = {a["name"]: i for i, a in enumerate(p["accounts"])}

    for line in budget["lines"]:
        idx = acct_map.get(line["deposit_account"])
        if idx is not None:
            p["accounts"][idx]["balance"] += line["deposit"]

    return p


def check_overdue(profile):
    """
    Return list of obligations whose due_date has passed and are not yet paid.
    """
    today = date.today()
    overdue = []
    for ob in profile.get("obligations", []):
        if ob.get("kind") == "allowance":
            continue  # allowance has no due date, never overdue
        try:
            due = datetime.strptime(ob["due_date"], "%Y-%m-%d").date()
            if due <= today and not ob.get("paid", False):
                overdue.append(ob)
        except Exception:
            pass
    return overdue


def mark_paid(profile, obligation_name):
    """
    Mark an obligation as paid:
    - Deduct target from its linked account
    - Goals (One-time): remove the obligation entirely — it's done
    - Expenses (recurring): advance the due date to the next period
    """
    import copy
    p = copy.deepcopy(profile)
    acct_map = {a["name"]: i for i, a in enumerate(p["accounts"])}

    for i, ob in enumerate(p["obligations"]):
        if ob["name"] == obligation_name:
            if ob.get("kind") == "allowance":
                break  # allowance never gets "paid" — skip
            # Deduct from account
            idx = acct_map.get(ob["deposit_account"])
            if idx is not None:
                p["accounts"][idx]["balance"] -= ob["target"]
            # Goals are one-time — remove them
            if ob.get("kind") == "goal" or ob.get("recurrence") == "One-time":
                p["obligations"].pop(i)
            elif ob.get("kind") == "running":
                # Running expenses roll forward AND reset target to 0
                # User will enter new balance next period
                rolled = advance_due_date(ob)
                rolled["target"] = 0.0
                p["obligations"][i] = rolled
            else:
                # Fixed expenses roll to next period
                p["obligations"][i] = advance_due_date(ob)
            break
    return p


# ══════════════════════════════════════════════════════════════════════════════
#  PROFILE CRUD
# ══════════════════════════════════════════════════════════════════════════════

def create_profile(uid, name, currency, pay_period,
                   obligations, accounts, theme=None):
    profile = {
        "uid":          uid,
        "name":         name,
        "currency":     currency,
        "pay_period":   pay_period,
        "created":      date.today().isoformat(),
        "obligations":  obligations,
        "accounts":     accounts,
        "paycheck_log": [],
        "theme":        theme or DEFAULT_THEME.copy(),
    }
    _write_xlsx(uid, profile)
    return uid


def load_profile(uid):
    path = _profile_path(uid)
    if not os.path.exists(path):
        raise FileNotFoundError(f"No profile found for ID: {uid}")
    return _read_wb(load_workbook(path, data_only=True), uid)


def save_profile(profile):
    _write_xlsx(profile["uid"], profile)


# ── Read XLSX → dict ──────────────────────────────────────────────────────────

def _read_wb(wb, uid):
    meta = wb["Meta"]
    profile = {
        "uid":          uid,
        "name":         meta["B2"].value,
        "currency":     meta["B3"].value,
        "pay_period":   meta["B4"].value,
        "created":      meta["B5"].value,
        "obligations":  [],
        "accounts":     [],
        "paycheck_log": [],
        "theme":        DEFAULT_THEME.copy(),
    }

    for row in meta.iter_rows(min_row=6, values_only=True):
        if row[0] and str(row[0]).startswith("theme_"):
            profile["theme"][str(row[0])[6:]] = row[1]

    if "Accounts" in wb.sheetnames:
        for row in wb["Accounts"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[0] != "TOTAL":
                profile["accounts"].append({
                    "name":    row[0],
                    "balance": row[1] or 0.0,
                    "type":    row[2] or "bank",
                })

    if "Obligations" in wb.sheetnames:
        for row in wb["Obligations"].iter_rows(min_row=2, values_only=True):
            if row[0]:
                profile["obligations"].append({
                    "name":            row[0],
                    "target":          row[1] or 0.0,
                    "recurrence":      row[2] or "Monthly",
                    "due_type":        row[3] or "Day of month",
                    "due_value":       str(row[4]) if row[4] else "1",
                    "due_date":        str(row[5]) if row[5] else "",
                    "deposit_account": row[6] or "",
                    "kind":            row[7] or "expense",
                    "paid":            str(row[8]).lower() == "true" if row[8] else False,
                })

    if "Paycheck Log" in wb.sheetnames:
        for row in wb["Paycheck Log"].iter_rows(min_row=2, values_only=True):
            if row[0]:
                profile["paycheck_log"].append({
                    "date":  str(row[0]),
                    "gross": row[1] or 0.0,
                    "net":   row[2] or 0.0,
                })

    return profile


# ── Write dict → XLSX ─────────────────────────────────────────────────────────

def _write_xlsx(uid, p):
    wb = Workbook(); wb.remove(wb.active)
    _sheet_meta(wb, p)
    _sheet_accounts(wb, p)
    _sheet_obligations(wb, p)
    _sheet_paycheck_log(wb, p)
    _sheet_budget_summary(wb, p)
    wb.save(_profile_path(uid))


def _sheet_meta(wb, p):
    ws = wb.create_sheet("Meta")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    _hdr(ws["A1"], "BudgetPal – User Profile"); ws.merge_cells("A1:B1")
    for label, value in [("Name", p["name"]), ("Currency", p["currency"]),
                          ("Pay Period", p["pay_period"]), ("Created", p["created"]),
                          ("User ID", p["uid"])]:
        r = ws.max_row + 1
        _nrm(ws.cell(r, 1), label); _inp(ws.cell(r, 2), value)
    for k, v in p.get("theme", {}).items():
        r = ws.max_row + 1
        _nrm(ws.cell(r, 1), f"theme_{k}"); _inp(ws.cell(r, 2), v)


def _sheet_accounts(wb, p):
    ws = wb.create_sheet("Accounts")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    _hdr(ws["A1"], "Account Name")
    _hdr(ws["B1"], "Balance")
    _hdr(ws["C1"], "Type")
    fmt = _currency_fmt(p["currency"])
    for i, a in enumerate(p.get("accounts", []), start=2):
        _inp(ws.cell(i, 1), a["name"])
        _inp(ws.cell(i, 2), a["balance"], num_fmt=fmt)
        _inp(ws.cell(i, 3), a.get("type", "bank"))
    last = len(p.get("accounts", [])) + 1
    tr = last + 1
    _sub(ws.cell(tr, 1), "TOTAL")
    ws.cell(tr, 2).value = f"=SUM(B2:B{last})" if last >= 2 else 0
    _style(ws.cell(tr, 2), font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
           border=BORDER, num_fmt=fmt)


def _sheet_obligations(wb, p):
    ws = wb.create_sheet("Obligations")
    cols = ["Name","Target","Recurrence","Due Type","Due Value",
            "Due Date","Deposit Account","Kind","Paid"]
    widths = [24, 14, 12, 14, 10, 14, 22, 10, 8]
    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[chr(64+ci)].width = w
        _hdr(ws.cell(1, ci), col)
    fmt = _currency_fmt(p["currency"])
    for i, ob in enumerate(p.get("obligations", []), start=2):
        _inp(ws.cell(i, 1), ob["name"])
        _inp(ws.cell(i, 2), ob["target"], num_fmt=fmt)
        _inp(ws.cell(i, 3), ob["recurrence"])
        _inp(ws.cell(i, 4), ob["due_type"])
        _inp(ws.cell(i, 5), ob["due_value"])
        _inp(ws.cell(i, 6), ob["due_date"])
        _inp(ws.cell(i, 7), ob["deposit_account"])
        _inp(ws.cell(i, 8), ob["kind"])
        _inp(ws.cell(i, 9), str(ob.get("paid", False)))


def _sheet_paycheck_log(wb, p):
    ws = wb.create_sheet("Paycheck Log")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    _hdr(ws["A1"], "Date"); _hdr(ws["B1"], "Gross"); _hdr(ws["C1"], "Net")
    fmt = _currency_fmt(p["currency"])
    for i, e in enumerate(p.get("paycheck_log", []), start=2):
        _nrm(ws.cell(i, 1), e["date"])
        _nrm(ws.cell(i, 2), e["gross"], num_fmt=fmt)
        _nrm(ws.cell(i, 3), e["net"],   num_fmt=fmt)


def _sheet_budget_summary(wb, p):
    ws = wb.create_sheet("Budget Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    _hdr(ws["A1"], "Budget Summary"); ws.merge_cells("A1:E1")

    fmt  = _currency_fmt(p["currency"])
    log  = p.get("paycheck_log", [])
    net  = log[-1]["net"] if log else 0.0
    if net == 0: return

    budget = build_budget(p, net)
    total_bal = sum(a["balance"] for a in p.get("accounts", []))

    for label, value in [("Total Bank Balance", total_bal),
                          ("Latest Net Pay", net),
                          ("Total Deposits Needed", budget["total_needed"]),
                          ("Shortfall", budget["shortfall"])]:
        r = ws.max_row + 1
        _sub(ws.cell(r, 1), label)
        _style(ws.cell(r, 2), font=SUBHEADER_FONT, fill=SUBHEADER_FILL,
               border=BORDER, num_fmt=fmt)
        ws.cell(r, 2).value = value

    ws.append([])
    r = ws.max_row + 1
    for ci, lbl in enumerate(["Obligation","Account","Target","Deposit","Spendable"], 1):
        _hdr(ws.cell(r, ci), lbl)
    for line in budget["lines"]:
        r = ws.max_row + 1
        _nrm(ws.cell(r, 1), line["name"])
        _nrm(ws.cell(r, 2), line["deposit_account"])
        _nrm(ws.cell(r, 3), line["target"],    num_fmt=fmt)
        _nrm(ws.cell(r, 4), line["deposit"],   num_fmt=fmt)
        _nrm(ws.cell(r, 5), line["spendable"], num_fmt=fmt)
